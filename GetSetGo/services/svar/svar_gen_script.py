import os
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

CATEGORY_MAP = {
    'NONE': 'SVAR_CAT_NONE',
    'CONFIG': 'SVAR_CAT_CONFIG',
    'RUNTIME': 'SVAR_CAT_RUNTIME',
    'EVENT': 'SVAR_CAT_EVENT',
}

# ---------------- TYPE MAP ----------------
TYPE_MAP = {
    'UINT8':  ('SVAR_TYPE_UINT8',  'u8', 1),
    'UINT16': ('SVAR_TYPE_UINT16', 'u16', 2),
    'UINT32': ('SVAR_TYPE_UINT32', 'u32', 4),
    'UINT64': ('SVAR_TYPE_UINT64', 'u64', 8),
    'INT8':   ('SVAR_TYPE_INT8',   'i8', 1),
    'INT16':  ('SVAR_TYPE_INT16',  'i16', 2),
    'INT32':  ('SVAR_TYPE_INT32',  'i32', 4),
    'INT64':  ('SVAR_TYPE_INT64',  'i64', 8),
    'FLOAT':  ('SVAR_TYPE_FLOAT',  'f',  4),
    'BOOL':   ('SVAR_TYPE_BOOL',   'b',  1),
    'STRING': ('SVAR_TYPE_STRING', 'str', None),
}

def resolve_type(type_str):
    if not type_str:
        return TYPE_MAP['UINT32']
    return TYPE_MAP.get(type_str.upper(), TYPE_MAP['UINT32'])

# ANSI color codes
class Colors:
    RESET = '\033[0m'
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'

def log_error(message):
    print(f"{Colors.RED}[ERROR]{Colors.RESET} {message}")

def log_warning(message):
    print(f"{Colors.YELLOW}[WARNING]{Colors.RESET} {message}")

def log_pass(message):
    print(f"{Colors.GREEN}[PASS]{Colors.RESET} {message}")

def log_info(message):
    print(f"{Colors.BLUE}[INFO]{Colors.RESET} {message}")

def log_debug(message):
    print(f"{Colors.CYAN}[DEBUG]{Colors.RESET} {message}")

def log_debug_block(lines):
    if not lines:
        return

    prefix = f"{Colors.CYAN}[DEBUG]{Colors.RESET} "
    print(prefix + lines[0])
    for line in lines[1:]:
        print(' ' * 16 + line)


def get_type_size(svar_type):
    """
    Return the size in bytes for a given svar_type_t.
    """
    size_map = {
        'SVAR_TYPE_INT8': 1,
        'SVAR_TYPE_INT16': 2,
        'SVAR_TYPE_INT32': 4,
        'SVAR_TYPE_INT64': 8,
        'SVAR_TYPE_UINT8': 1,
        'SVAR_TYPE_UINT16': 2,
        'SVAR_TYPE_UINT32': 4,
        'SVAR_TYPE_UINT64': 8,
        'SVAR_TYPE_FLOAT': 4,
        'SVAR_TYPE_BOOL': 1,
        'SVAR_TYPE_CHAR': 1,
        'SVAR_TYPE_STRING': 256,  # Default string size
        'SVAR_TYPE_GROUP': 4,
    }
    return size_map.get(svar_type, 4)


def get_union_field_and_cast(svar_type, value):
    """
    Return the union field name and properly casted value for the given svar_type.
    e.g., (SVAR_TYPE_UINT8, 50) -> ("u8", "50")
    """
    if not isinstance(value, (int, float, bool, str)):
        value = 0
    
    type_field_map = {
        'SVAR_TYPE_INT8': 'i8',
        'SVAR_TYPE_INT16': 'i16',
        'SVAR_TYPE_INT32': 'i32',
        'SVAR_TYPE_INT64': 'i64',
        'SVAR_TYPE_UINT8': 'u8',
        'SVAR_TYPE_UINT16': 'u16',
        'SVAR_TYPE_UINT32': 'u32',
        'SVAR_TYPE_UINT64': 'u64',
        'SVAR_TYPE_FLOAT': 'f',
        'SVAR_TYPE_BOOL': 'b',
        'SVAR_TYPE_CHAR': 'c',
        'SVAR_TYPE_STRING': 'str',
    }
    
    field = type_field_map.get(svar_type, 'u32')
    return field


def get_excel_field(obj, field_names):
    """
    Try to get a value from object using any of the field_names.
    """
    if not isinstance(obj, dict):
        return None
    
    for name in field_names:
        value = obj.get(name)
        if value is not None:
            if isinstance(value, str):
                value = value.strip()
                if value:
                    return value
            else:
                return value
    
    return None


def get_excel_raw_field(obj, field_names):
    """
    Return a raw cell value using any of the field_names.
    Empty strings are returned as empty strings.
    """
    if not isinstance(obj, dict):
        return None

    for name in field_names:
        if name in obj:
            value = obj.get(name)
            if isinstance(value, str):
                return value.strip()
            return value

    return None


def sanitize_enum_name(name):
    if not name or not isinstance(name, str):
        return 'UNKNOWN'

    cleaned = ''.join(c if c.isalnum() else '_' for c in name)
    cleaned = cleaned.strip('_')
    if cleaned and cleaned[0].isdigit():
        cleaned = '_' + cleaned
    return cleaned.upper()


def get_variable_name(obj):
    """Resolve the enum/source variable name from the Name column."""
    if not isinstance(obj, dict):
        return ''

    value = obj.get('Name')
    if isinstance(value, str):
        return value.strip()
    if value is not None:
        return str(value).strip()

    return ''


def get_display_name(obj):
    """Resolve the string used for system_variable_t.name from DisplayName."""
    value = get_excel_raw_field(obj, ['DisplayName', 'displayName', 'displayname'])
    if value is None:
        return ''

    return str(value).strip()


def describe_row(obj):
    row_num = obj.get('__row_num', '?') if isinstance(obj, dict) else '?'
    name = get_variable_name(obj)
    if name:
        return f"variables row {row_num}, Name='{name}'"

    return f"variables row {row_num}"


def parse_int_value(value, context, default=0):
    if value is None:
        return default

    if isinstance(value, str):
        value = value.strip()
        if value == '':
            return default

    try:
        return int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{context}: expected integer, got {value!r}")


def parse_int_field(obj, field_names, default=0):
    value = get_excel_raw_field(obj, field_names)
    field_label = '/'.join(field_names)
    return parse_int_value(value, f"{describe_row(obj)} field '{field_label}'", default)


def parse_config_int(config_data, key, default=0):
    return parse_int_value(config_data.get(key), f"config field '{key}'", default)


def get_svar_comment(obj):
    """Resolve optional enum comment text from the spreadsheet."""
    comment = get_excel_field(obj, ['comment', 'Comment', 'COMMENT'])
    if comment is None:
        return ''

    return str(comment).replace('\r', ' ').replace('\n', ' ').strip()


def c_escape_string(value):
    """Escape text for use inside a C string literal."""
    return str(value).replace('\\', '\\\\').replace('"', '\\"')


def format_string_default(value):
    """Return spreadsheet Default text for string SVARs."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value

    return str(value)


def format_svar_name_initializer(name):
    """Return the C initializer for system_variable_t.name."""
    if name == '' or name == 'NULL':
        return 'NULL'

    return f'"{c_escape_string(name)}"'


def truncate_name(name, max_length):
    """Truncate name to max_length, preserving prefix if it starts with SVAR_"""
    if not name or max_length <= 0:
        return ''
    
    # Check if name starts with SVAR_ prefix
    prefix = ''
    base_name = name
    if name.upper().startswith('SVAR_'):
        prefix = 'SVAR_'
        base_name = name[5:]  # Remove 'SVAR_' prefix
    
    # Calculate available length for base name
    available_len = max_length - len(prefix)
    
    if len(base_name) <= available_len:
        return name
    
    # Truncate base name and reassemble
    truncated = prefix + base_name[:available_len]
    return truncated

class SVARDatabase:
    def __init__(self):
        self.setup_info = {}
        self.setup_config = {}
        self.configurations = defaultdict(list)
        self.current_section = None
        self.current_section_type = None

def read_database(excel_file_path):
    log_info(f"Reading database from: {excel_file_path}")

    config_data = {}
    product_name = 'Default'
    svar_offset = 0
    svar_name_max_length = 16  # Default value
    try:
        workbook = load_workbook(excel_file_path, data_only=True)

        # Read config sheet first
        if 'config' in workbook.sheetnames:
            config_sheet = workbook['config']
            config_rows = list(config_sheet.iter_rows(values_only=True))
            for row in config_rows:
                if row and len(row) >= 2:
                    key = str(row[0]).strip() if row[0] else ""
                    value = str(row[1]).strip() if row[1] else ""
                    if key:
                        config_data[key] = value
                    # Extract NAME field for product naming
                    if key.upper() == 'NAME':
                        product_name = value if value else 'Default'
                    # Extract SVAR_OFFSET field
                    if key.upper() == 'SVAR_OFFSET':
                        svar_offset = parse_int_value(value, "config field 'SVAR_OFFSET'", 0)
                    # Extract SVAR_NAME_MAX_LENGTH field (from next cell)
                    if key.upper() == 'MAX_NAME_LEN':
                        svar_name_max_length = parse_int_value(value, "config field 'MAX_NAME_LEN'", 8)
                        # if len(row) >= 3 and row[2]:
                            # svar_name_max_length = int(row[2]) if row[2] else 16
            log_debug(f"Read config: {config_data}")
            log_info(f"Product name: {product_name}, SVAR_OFFSET: {svar_offset}")

        if 'variables' not in workbook.sheetnames:
            log_error("Sheet 'variables' not found")
            return None, config_data, 'Default', 0, 16

        sheet = workbook['variables']

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], config_data, product_name, svar_offset, svar_name_max_length

        # First row = header
        headers = [str(h).strip() if h else "" for h in rows[0]]

        data = []

        for row_num, row in enumerate(rows[1:], start=2):
            obj = {}

            for i, header in enumerate(headers):
                if not header:
                    continue

                value = row[i] if i < len(row) else None

                # Normalize strings
                if isinstance(value, str):
                    value = value.strip()

                obj[header] = value

            obj['__row_num'] = row_num
            if not get_variable_name(obj):
                log_info(f"Stopping variable parse at row {row_num}: Name column is empty")
                break

            data.append(obj)

        # Add incrementing id to each object, starting from 1
        for i, obj in enumerate(data, start=0):
            obj['id'] = i

        workbook.close()
        log_pass(f"Parsed {len(data)} variables")
        
        # Debug: Print parsed objects in readable format
        debug_lines = ["Parsed objects:"]
        for i, obj in enumerate(data[:5]):  # Show first 5 objects
            debug_lines.append(f"Object {i+1}:")
            for key, value in obj.items():
                debug_lines.append(f"{key}: {value}")
            if i < min(len(data), 5) - 1:
                debug_lines.append("")
        if len(data) > 5:
            debug_lines.append(f"... and {len(data) - 5} more objects")
        log_debug_block(debug_lines)

        return data, config_data, product_name, svar_offset, svar_name_max_length

    except FileNotFoundError:
        log_error(f"Excel file not found: {excel_file_path}")
        return None, config_data, 'Default', 0, 16
    except Exception as e:
        log_error(f"Error reading Excel file: {str(e)}")
        return None, config_data, 'Default', 0, 16

def create_app_header(data, output_dir, config_data=None, product_name='Default', svar_offset=0, svar_name_max_length=16):
    """
    Create a svar_<NAME>.h file with NVM address allocations
    """
    if config_data is None:
        config_data = {}
    
    # Sanitize product name for use in C identifiers
    product_prefix = sanitize_enum_name(product_name)
    
    os.makedirs(output_dir, exist_ok=True)
    header_file = os.path.join(output_dir, f'svar_{product_prefix.lower()}.h')

    try:
        with open(header_file, 'w') as f:
            config_start  = parse_config_int(config_data, 'NVM_START_ADD_CONFIG', 0)
            runtime_start = parse_config_int(config_data, 'NVM_START_ADD_RUNTIME', 0)
            event_start   = parse_config_int(config_data, 'NVM_START_ADD_EVENT', 0)
            nvm_end       = parse_config_int(config_data, 'NVM_END_ADD', 0)

            config_counter  = config_start
            runtime_counter = runtime_start
            event_counter   = event_start

            f.write(f'#ifndef SVAR_{product_prefix}_H_\n')
            f.write(f'#define SVAR_{product_prefix}_H_\n')
            f.write('#include "gsg_config.h"\n')
            f.write('#include "services/svar/svar_internal.h"\n\n')

            f.write('/* Configurations */\n')
            # f.write('#define SVAR_MAX_VARIABLES      100\n')
            f.write(f'#define SVAR_NAME_MAX_LENGTH    {svar_name_max_length}\n\n')

            # SVAR_OFFSET define
            f.write('/* SVAR Offset */\n')
            f.write(f'#define SVAR_OFFSET_{product_prefix}   {svar_offset}\n\n')

            # Create enum with variable IDs
            f.write('/* Variable ID Enum */\n')
            f.write('typedef enum {\n')            
            for obj in data:
                var_name = get_variable_name(obj)
                if var_name:
                    enum_name = sanitize_enum_name(var_name.replace('SVAR_', ''))
                    comment = get_svar_comment(obj)
                    comment_suffix = f' // {comment}' if comment else ' //'
                    f.write(f'    SVAR_{product_prefix}_{enum_name:<25} = SVAR_OFFSET_{product_prefix} + {obj["id"]},\t\t{comment_suffix}\n')
            # f.write(f'    __SVAR_{product_prefix}_MAX_ID = SVAR_OFFSET_{product_prefix} + {len(data) + 1}\n')
            f.write(f'    __SVAR_{product_prefix}_MAX_ID{"":<{34 - len(f"__SVAR_{product_prefix}_MAX_ID")}} = SVAR_OFFSET_{product_prefix} + {len(data)}\n')
            f.write(f'}} svar_{product_prefix}_id_t;\n\n')
            
            # Generate NVM address defines for persistent variables
            f.write('/* NVM Address Allocation */\n')
            f.write(f'#define NVM_ADD_SVAR_{product_prefix}_CONFIG_BASE   {config_start}\n')
            f.write(f'#define NVM_ADD_SVAR_{product_prefix}_RUNTIME_BASE  {runtime_start}\n')
            f.write(f'#define NVM_ADD_SVAR_{product_prefix}_EVENT_BASE    {event_start}\n\n')
            
            persistent_vars = []
            for obj in data:
                category_str = get_excel_field(obj, ['Category']) or 'NONE'
                category_enum = CATEGORY_MAP.get(category_str.upper(), 'SVAR_CAT_NONE')
                var_name = get_variable_name(obj)
                if var_name:
                    # Truncate variable name to svar_name_max_length
                    var_name = truncate_name(var_name, svar_name_max_length)
                    persistent = get_excel_field(obj, ['Persistent', 'persistent'])
                    if persistent and str(persistent).lower() in ['yes', 'true', '1']:
                        type_str = get_excel_field(obj, ['Type'])
                        svar_type, field, size = resolve_type(type_str)
                        type_size = size
                        define_name = sanitize_enum_name(var_name.replace('SVAR_', ''))
                        if svar_type == 'SVAR_TYPE_STRING':
                            type_size = parse_int_field(obj, ['Max'], 0)

                        define_name = sanitize_enum_name(var_name.replace('SVAR_', ''))

                        if category_enum == 'SVAR_CAT_CONFIG':
                            offset = config_counter - config_start
                            f.write(f'#define NVM_ADD_SVAR_{product_prefix}_{define_name:<30} (NVM_ADD_SVAR_{product_prefix}_CONFIG_BASE + {offset})\n')
                            config_counter += type_size

                        elif category_enum == 'SVAR_CAT_RUNTIME':
                            offset = runtime_counter - runtime_start
                            f.write(f'#define NVM_ADD_SVAR_{product_prefix}_{define_name:<30} (NVM_ADD_SVAR_{product_prefix}_RUNTIME_BASE + {offset})\n')
                            runtime_counter += type_size

                        elif category_enum == 'SVAR_CAT_EVENT':
                            offset = event_counter - event_start
                            f.write(f'#define NVM_ADD_SVAR_{product_prefix}_{define_name:<30} (NVM_ADD_SVAR_{product_prefix}_EVENT_BASE + {offset})\n')
                            event_counter += type_size
            
            if persistent_vars:
                f.write(f'\n#define NVM_ADD_SVAR_{product_prefix}_CONFIG_END   {config_counter}\n')
                f.write(f'#define NVM_ADD_SVAR_{product_prefix}_RUNTIME_END {runtime_counter}\n')
                f.write(f'#define NVM_ADD_SVAR_{product_prefix}_EVENT_END   {event_counter}\n\n')
            else:
                f.write('\n')

            # Extern module declaration
            f.write(f'/* SVAR Module */\n')
            f.write(f'extern svar_module_t svar_module_{product_prefix};\n\n')

            f.write(f'#endif /* SVAR_{product_prefix}_H_ */\n')

        log_pass(f"Created {header_file}")
        log_info(f"Allocated {len(persistent_vars)} persistent variables in NVM")
    except Exception as e:
        log_error(f"Failed to create {header_file}: {str(e)}")


def create_app_table(data, output_dir, config_data=None, product_name='Default', svar_offset=0, svar_name_max_length=16):
    # Sanitize product name for use in C identifiers
    product_prefix = sanitize_enum_name(product_name)
    
    os.makedirs(output_dir, exist_ok=True)
    source_file = os.path.join(output_dir, f'svar_{product_prefix.lower()}.c')

    try:
        # -------- NAME → ID MAP --------
        name_to_id = {}
        for obj in data:
            name = get_variable_name(obj)
            if name:
                name_to_id[name] = obj['id']

        # -------- STRING + NVM PREPASS --------
        string_buffers = []
        nvm_map = {}
        config_start  = parse_config_int(config_data, 'NVM_START_ADD_CONFIG', 0)
        runtime_start = parse_config_int(config_data, 'NVM_START_ADD_RUNTIME', 0)
        event_start   = parse_config_int(config_data, 'NVM_START_ADD_EVENT', 0)
        nvm_end       = parse_config_int(config_data, 'NVM_END_ADD', 0)

        config_counter  = config_start
        runtime_counter = runtime_start
        event_counter   = event_start

        for obj in data:
            var_name = get_variable_name(obj)
            if not var_name:
                continue

            type_str = get_excel_field(obj, ['Type'])
            svar_type, field, size = resolve_type(type_str)

            persistent = get_excel_field(obj, ['Persistent'])
            is_persistent = str(persistent).lower() in ['yes', 'true', '1']

            if svar_type == 'SVAR_TYPE_STRING':
                max_len = parse_int_field(obj, ['Max'], 0)
                string_buffers.append((obj['id'], max_len))
                size = max_len

            category_str = get_excel_field(obj, ['Category']) or 'NONE'
            category_enum = CATEGORY_MAP.get(category_str.upper(), 'SVAR_CAT_NONE')
            if is_persistent and size:
                macro = f"NVM_ADD_SVAR_{product_prefix}_{sanitize_enum_name(var_name)}"
                if category_enum == 'SVAR_CAT_CONFIG':
                    addr = config_counter
                    config_counter += size
                elif category_enum == 'SVAR_CAT_RUNTIME':
                    addr = runtime_counter
                    runtime_counter += size
                elif category_enum == 'SVAR_CAT_EVENT':
                    addr = event_counter
                    event_counter += size
                else:
                    addr = 0  # no allocation
                nvm_map[obj['id']] = macro
        # -------- NVM RANGE CHECK --------
        if config_counter > nvm_end:
            log_error("CONFIG NVM overflow")
            return

        if runtime_counter > nvm_end:
            log_error("RUNTIME NVM overflow")
            return

        if event_counter > nvm_end:
            log_error("EVENT NVM overflow")
            return
        # -------- NVM OVERLAP CHECK --------
        ranges = [
            ('CONFIG', config_start, config_counter),
            ('RUNTIME', runtime_start, runtime_counter),
            ('EVENT', event_start, event_counter),
        ]

        for i in range(len(ranges)):
            for j in range(i + 1, len(ranges)):
                name1, s1, e1 = ranges[i]
                name2, s2, e2 = ranges[j]

                if not (e1 <= s2 or e2 <= s1):
                    log_error(f"NVM overlap between {name1} and {name2}")
                    return

        # -------- WRITE FILE --------
        with open(source_file, 'w') as f:
            f.write('#include "gsg_config.h"\n')
            f.write('#ifdef GSG_USE_SVAR\n')
            f.write('#if (GSG_USE_SVAR == GSG_ENABLE)\n')

            f.write(f'#include "svar_{product_prefix.lower()}.h"\n\n')
            # f.write('#include "svar.h"\n')
            # f.write('#include "svar_internal.h"\n\n')

            # STRING BUFFERS
            for sid, size in string_buffers:
                f.write(f'static char svar_{sid}_buf[{size}];\n')
            if string_buffers:
                f.write('\n')

            f.write('static system_variable_t svar_table[] = {\n')

            # Create index counter starting from 0
            table_index = 0
            for obj in data:
                raw_var_name = get_variable_name(obj)
                if not raw_var_name:
                    continue

                # Truncate variable name to svar_name_max_length
                var_name = truncate_name(raw_var_name, svar_name_max_length)

                type_str = get_excel_field(obj, ['Type'])
                svar_type, field, size = resolve_type(type_str)

                value = get_excel_field(obj, ['Default'])
                min_val = get_excel_field(obj, ['Min'])
                max_val = get_excel_field(obj, ['Max'])

                set_cb = get_excel_field(obj, ['WriteCallback'])
                get_cb = get_excel_field(obj, ['readCallback'])

                parent_name = get_excel_field(obj, ['Parent'])
                parent_id = name_to_id.get(parent_name, 0)

                persistent = get_excel_field(obj, ['Persistent'])
                is_persistent = str(persistent).lower() in ['yes', 'true', '1']

                f.write(f'\t[{table_index}] = {{\n')
                f.write(f'\t\t.id = SVAR_OFFSET_{product_prefix} + {obj["id"]},\n')
                display_name = get_display_name(obj)
                name_initializer = format_svar_name_initializer(display_name)
                f.write(f'\t\t.name = {name_initializer},\n')
                f.write(f'\t\t.type = {svar_type},\n')
                f.write(f'\t\t.parent = {parent_id},\n')

                category_str = get_excel_field(obj, ['Category']) or 'NONE'
                category_enum = CATEGORY_MAP.get(category_str.upper(), 'SVAR_CAT_NONE')
                f.write(f'\t\t.category = {category_enum},\n')

                if svar_type == 'SVAR_TYPE_STRING':
                    max_len = parse_int_field(obj, ['Max'], 0)
                    f.write(f'\t\t.value = {{ .str = svar_{obj["id"]}_buf }},\n')
                    f.write(f'\t\t.def   = {{ .str = "{c_escape_string(format_string_default(value))}" }},\n')
                    f.write(f'\t\t.min = {{ .u16 = 0 }},')
                    f.write(f'\t\t.max = {{ .u16 = {max_len} }},\n')
                    # f.write(f'\t\t.maxLen = {max_len},\n')
                else:
                    value = parse_int_field(obj, ['Default'], 0)
                    min_val = parse_int_field(obj, ['Min'], 0)
                    max_val = parse_int_field(obj, ['Max'], 0)

                    f.write(f'\t\t.value = {{.{field} = {value}}},')
                    f.write(f'\t\t.def   = {{.{field} = {value}}},\n')
                    f.write(f'\t\t.min   = {{.{field} = {min_val}}},')
                    f.write(f'\t\t.max   = {{.{field} = {max_val}}},\n')

                f.write(f'\t\t.nvmAddr = {nvm_map.get(obj["id"], 0)},\n')

                f.write(f'\t\t.flags = {{\n')
                f.write(f'\t\t\t.persistent = {1 if is_persistent else 0},\n')
                f.write(f'\t\t}},\n')

                f.write(f'\t\t.setCb = {"NULL" if not set_cb else set_cb},\n')
                f.write(f'\t\t.getCb = {"NULL" if not get_cb else get_cb},\n')

                f.write('\t},\n')
                
                table_index += 1

            f.write('};\n\n')

            # Generate svar_module_t struct
            f.write(f'svar_module_t svar_module_{product_prefix} = {{\n')
            f.write(f'    .table = svar_table,\n')
            f.write(f'    .count = {len(data)},\n')
            f.write(f'    .varOffset = SVAR_OFFSET_{product_prefix}\n')
            f.write('};\n\n')

            f.write('#endif\n#endif\n')

        log_pass(f"Created {source_file}")

    except Exception as e:
        log_error(f"Failed to create {source_file}: {str(e)}")


def create_app_user(data, output_dir, product_name='Default'):
    """
    Create a svar_user.h file
    """
    os.makedirs(output_dir, exist_ok=True)
    user_header_file = os.path.join(output_dir, 'svar_user.h')

    # Sanitize product name for use in C identifiers
    product_prefix = sanitize_enum_name(product_name)

    try:
        with open(user_header_file, 'w') as f:
            f.write('#ifndef SVAR_USER_H_\n')
            f.write('#define SVAR_USER_H_\n\n')
            f.write('#include "svar_add.h"\n\n')

            f.write('#define SVAR_SEGMENT_COUNT      5\n')
            f.write('#define SEGMENT1_SIZE           100\n\n')

            f.write('typedef enum segmentList_t{\n')
            f.write('    RAM_ONLY,\n')
            f.write('    EEPROM_ONLY,\n')
            f.write('    SHADOWED_EEPROM\n')
            f.write('} segmentList_t;\n\n')

            # Add variable declarations
            f.write('/* Variable Declarations */\n')
            for obj in data:
                var_name = get_variable_name(obj)
                if var_name:
                    sanitized_name = sanitize_enum_name(var_name)
                    f.write(f'extern const system_variable_t {sanitized_name};\n')

            f.write('\n#endif /* SVAR_USER_H_ */\n')

        log_pass(f"Created {user_header_file}")
    except Exception as e:
        log_error(f"Failed to create {user_header_file}: {str(e)}")

def select_excel_file():
    """
    Open a file selection dialog for Excel files
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(
        title="Select SVAR Excel file",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    return file_path

# Example usage
if __name__ == "__main__":
    log_info("SVAR Code Generator Started")

    # Open file selection dialog
    excel_file = select_excel_file()

    if not excel_file:
        log_warning("No file selected. Exiting.")
        exit(1)

    log_info(f"Selected file: {excel_file}")

    # Read the database
    db, config_data, product_name, svar_offset, svar_name_max_length = read_database(excel_file)

    if db is None:
        log_error("Failed to read database. Exiting.")
        exit(1)

    log_info(f"Product name from config: {product_name}, SVAR_OFFSET: {svar_offset}")

    # Create output directory next to the Excel file
    excel_dir = os.path.dirname(excel_file)
    output_dir = os.path.join(excel_dir, f"svar_{product_name}")
    log_info(f"Creating output directory: {output_dir}")

    try:
        os.makedirs(output_dir, exist_ok=True)
        log_pass(f"Output directory created: {output_dir}")
    except Exception as e:
        log_error(f"Failed to create output directory: {str(e)}")
        exit(1)

    # Generate files
    create_app_header(db, output_dir, config_data, product_name, svar_offset, svar_name_max_length)
    create_app_table(db, output_dir, config_data, product_name, svar_offset, svar_name_max_length)
    # create_app_user(db, output_dir, product_name)

    log_pass("All files generated successfully!")
    log_info(f"Files created in: {output_dir}")
